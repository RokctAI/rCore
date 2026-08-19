# Copyright (c) 2026 RokctAI
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""Regenerate known_schools.json from the DBE EMIS Schools Masterlist.

Source
------
Department of Basic Education, EMIS "Schools Masterlist Data" page:
    https://www.education.gov.za/Programmes/EMIS/EMISDownloads.aspx
National masterlist file (all 9 provinces, public + independent, incl.
special-needs centres), served as National.xlsx:
    https://www.education.gov.za/LinkClick.aspx?fileticket=cF64WfdfGD0%3d&tabid=466&portalid=0&mid=14616&forcedownload=true

Usage
-----
    pip install openpyxl
    python3 transform_masterlist.py /path/to/National.xlsx

Writes known_schools.json next to this script. The transform is purely
mechanical — every record comes verbatim from a masterlist row (NatEmis and
school name are never invented or edited beyond whitespace collapsing);
province codes and phase labels are mapped through the fixed tables below.
Rows are deduped by NatEmis and sorted by (province, name, natemis) for
stable diffs.
"""

import json
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

PROVINCES = {
    "EC": "Eastern Cape",
    "FS": "Free State",
    "GT": "Gauteng",
    "GP": "Gauteng",
    "KZN": "KwaZulu-Natal",
    "LP": "Limpopo",
    "MP": "Mpumalanga",
    "NC": "Northern Cape",
    "NW": "North West",
    "WC": "Western Cape",
}

PHASES = {
    "PRIMARY SCHOOL": "primary",
    "SECONDARY SCHOOL": "secondary",
    "COMBINED SCHOOL": "combined",
    "INTERMEDIATE SCHOOL": "intermediate",
    "SPECIAL NEEDS EDUCATION SCHOOL": "special_needs",
    "SCHOOL OF SKILLS": "school_of_skills",
    "HOSPITAL SCHOOL": "hospital",
    "ECD": "ecd",
}

PLACEHOLDERS = {"", "99", "0", "UNKNOWN", "N/A", "NONE"}

README = [
    "Seed data for the school-capture feature (see the Dart mirror in",
    "lms/dart/lib/src/common/domain/models/known_schools.dart — keep the two",
    "in sync until the backend serves this list). Curricula with a seed list",
    "here are the finite, knowable ones; unbounded curricula (US Common",
    "Core, Cambridge, ...) deliberately have NO seed list — their suggestion",
    "pool is the schools students actually enter, accumulated server-side.",
    "CAPS holds every school in the DBE EMIS national masterlist — public",
    "AND registered independent (see each record's 'sector') — because the",
    "masterlist does not say which exam body an independent school uses, so",
    "true IEB schools cannot be split out from it; IEB therefore stays an",
    "accumulate-server-side bucket. Records are verbatim masterlist rows",
    "(names keep the source's ALL-CAPS; server-side normalisation for",
    "display/matching is a handoff item). Regenerate with",
    "transform_masterlist.py — never hand-edit or fabricate entries.",
]


def clean(value):
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return None if text.upper() in PLACEHOLDERS else text


def main() -> None:
    xlsx = Path(sys.argv[1] if len(sys.argv) > 1 else "National.xlsx")
    out = Path(__file__).resolve().parent / "known_schools.json"

    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c) for c in next(rows)]
    ix = {h: i for i, h in enumerate(header)}

    schools = {}
    data_years = set()
    for row in rows:
        natemis = clean(row[ix["NatEmis"]])
        name = clean(row[ix["Official_Institution_Name"]])
        if not natemis or not name:
            continue
        if natemis in schools:
            continue  # dedupe by NatEmis, first occurrence wins
        data_years.add(clean(row[ix["DataYear"]]))
        record = {
            "natemis": natemis,
            "name": name,
            "province": PROVINCES[clean(row[ix["Province"]])],
            "sector": clean(row[ix["Sector"]]).lower(),
            "phase": PHASES[clean(row[ix["Phase_PED"]]).upper()],
        }
        town = clean(row[ix["Town_City"]])
        if town:
            record["town"] = town
        schools[natemis] = record

    ordered = sorted(
        schools.values(), key=lambda s: (s["province"], s["name"], s["natemis"])
    )

    doc_head = {
        "_readme": README,
        "_source": {
            "name": "DBE EMIS Schools Masterlist (National), DataYear "
            + "/".join(sorted(y for y in data_years if y)),
            "masterlist_page": "https://www.education.gov.za/Programmes/EMIS/EMISDownloads.aspx",
            "file_url": "https://www.education.gov.za/LinkClick.aspx?fileticket=cF64WfdfGD0%3d&tabid=466&portalid=0&mid=14616&forcedownload=true",
            "retrieved": date.today().isoformat(),
            "school_count": len(ordered),
        },
    }

    # One school per line: keeps the file diff-able without ballooning it.
    with out.open("w", encoding="utf-8") as f:
        head = json.dumps(doc_head, ensure_ascii=False, indent=2)
        f.write(head[:-2].rstrip() + ",\n")  # strip closing "\n}" and reopen
        f.write('  "curricula": {\n    "CAPS": [\n')
        for i, rec in enumerate(ordered):
            line = json.dumps(rec, ensure_ascii=False)
            f.write("      " + line + (",\n" if i < len(ordered) - 1 else "\n"))
        f.write('    ],\n    "IEB": []\n  }\n}\n')

    print(f"wrote {out} with {len(ordered)} schools")


if __name__ == "__main__":
    main()
